"""Per-tour layout ("style") overrides on top of the shared xlsx template.

Some tours need a one-off manual layout tweak (e.g. a cell's fill/border
adjusted for that tour's particular content) beyond what the shared
template provides. This module lets that fix be captured once -- from a
user's manually-corrected copy of a generated workbook -- and reapplied
automatically every future time that tour's documents are generated.

Privacy: only cell *styling* (fill/border/font/number-format), column
widths, and merged-cell ranges are captured, keyed by cell reference.
Cell *values* are read only to align rows/columns; nothing derived from
booking/customer data is ever stored. See db.py's tour_style_overrides
table, which stores exactly the JSON this module produces, keyed by
tour_name (no booking data).
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from . import xlsx_lowlevel as xl
from .tour_workbook import SHEET1, SHEET2, TEMPLATE_PATH

SHEET_NAME_TO_PART = {
    "予約情報": SHEET1,
    "予約確認書(Oyster)": SHEET2,
}
STYLES_PART = "xl/styles.xml"


def _side_key(side) -> tuple[str | None, str | None]:
    if side is None:
        return (None, None)
    return (side.style, side.color.rgb if side.color else None)


def _style_key(cell) -> tuple:
    fill = cell.fill
    fill_key = (
        (fill.patternType, fill.fgColor.rgb if fill.fgColor else None) if fill else None
    )
    border = cell.border
    border_key = (
        (_side_key(border.left), _side_key(border.right), _side_key(border.top), _side_key(border.bottom))
        if border
        else None
    )
    font = cell.font
    font_key = (
        (font.name, font.sz, font.bold, font.italic, font.color.rgb if font.color else None) if font else None
    )
    return (fill_key, border_key, font_key, cell.number_format)


def capture_style_diff(fixed_path: Path) -> dict:
    """Diff `fixed_path` (a user's manually-corrected copy of a generated
    tour workbook) against the shared base template, cell by cell, and
    return only what differs: styled cells (by reference, no values),
    changed column widths, and any newly-added merged-cell ranges.
    """
    base_wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
    fixed_wb = openpyxl.load_workbook(Path(fixed_path), data_only=False)
    try:
        sheets: dict[str, dict] = {}
        for sheet_name in SHEET_NAME_TO_PART:
            if sheet_name not in base_wb.sheetnames or sheet_name not in fixed_wb.sheetnames:
                continue
            base_ws = base_wb[sheet_name]
            fixed_ws = fixed_wb[sheet_name]

            cell_refs = []
            for row in range(1, base_ws.max_row + 1):
                for col in range(1, base_ws.max_column + 1):
                    base_cell = base_ws.cell(row=row, column=col)
                    fixed_cell = fixed_ws.cell(row=row, column=col)
                    if _style_key(base_cell) != _style_key(fixed_cell):
                        cell_refs.append(fixed_cell.coordinate)

            col_widths = {}
            for letter, dim in fixed_ws.column_dimensions.items():
                if not dim.width:
                    continue
                base_dim = base_ws.column_dimensions.get(letter)
                if base_dim is None or base_dim.width != dim.width:
                    col_widths[letter] = dim.width

            base_merges = {str(r) for r in base_ws.merged_cells.ranges}
            fixed_merges = {str(r) for r in fixed_ws.merged_cells.ranges}
            extra_merges = sorted(fixed_merges - base_merges)

            if cell_refs or col_widths or extra_merges:
                sheets[sheet_name] = {
                    "cell_refs": cell_refs,
                    "col_widths": col_widths,
                    "extra_merges": extra_merges,
                }
        return {"sheets": sheets}
    finally:
        base_wb.close()
        fixed_wb.close()


def _tag_items(inner: str, item_tag: str) -> list[str]:
    pattern = re.compile(rf"<{item_tag}\b[^>]*/>|<{item_tag}\b[^>]*>.*?</{item_tag}>", re.S)
    return pattern.findall(inner)


def _section_items(styles_xml: str, tag: str, item_tag: str) -> list[str]:
    match = re.search(rf"<{tag}\b[^>]*>(.*?)</{tag}>", styles_xml, re.S)
    if not match:
        return []
    return _tag_items(match.group(1), item_tag)


def _numfmts(styles_xml: str) -> dict[int, str]:
    match = re.search(r"<numFmts\b[^>]*>(.*?)</numFmts>", styles_xml, re.S)
    if not match:
        return {}
    result = {}
    for entry in re.findall(r"<numFmt\b[^>]*/>", match.group(1)):
        id_match = re.search(r'numFmtId="(\d+)"', entry)
        if id_match:
            result[int(id_match.group(1))] = entry
    return result


_XF_ATTR_RE = re.compile(r'(\w+)="([^"]*)"')


def _resolve_xf(xf_xml: str, fonts: list[str], fills: list[str], borders: list[str], numfmts: dict[int, str]) -> dict:
    open_match = re.match(r"<xf([^>]*?)(/?)>", xf_xml)
    attrs = dict(_XF_ATTR_RE.findall(open_match.group(1)))
    self_closing = open_match.group(2) == "/"
    font_id = int(attrs.pop("fontId", 0))
    fill_id = int(attrs.pop("fillId", 0))
    border_id = int(attrs.pop("borderId", 0))
    numfmt_id = int(attrs.pop("numFmtId", 0))
    attrs.pop("xfId", None)
    if self_closing:
        extra_xml = ""
    else:
        inner_match = re.search(r"<xf[^>]*>(.*)</xf>", xf_xml, re.S)
        extra_xml = inner_match.group(1) if inner_match else ""
    return {
        "font_xml": fonts[font_id] if font_id < len(fonts) else (fonts[0] if fonts else "<font/>"),
        "fill_xml": fills[fill_id] if fill_id < len(fills) else (fills[0] if fills else "<fill/>"),
        "border_xml": borders[border_id] if border_id < len(borders) else (borders[0] if borders else "<border/>"),
        "numfmt_id": numfmt_id,
        "numfmt_xml": numfmts.get(numfmt_id),
        "xf_attrs": attrs,
        "xf_extra_xml": extra_xml,
    }


def _cell_style_index(sheet_xml: str, ref: str) -> int | None:
    match = re.search(rf'<c r="{re.escape(ref)}"[^>]*>', sheet_xml)
    if not match:
        return None
    s_match = re.search(r's="(\d+)"', match.group(0))
    return int(s_match.group(1)) if s_match else 0


def extract_cell_styles(fixed_path: Path, diff: dict) -> dict:
    """Resolve each diffed cell reference to its actual style definition
    (raw font/fill/border/numFmt XML from `fixed_path`), so the diff can be
    replayed onto a different workbook later regardless of that workbook's
    own style-table indices.
    """
    parts = xl.load_zip_parts(Path(fixed_path))
    styles_xml = parts[STYLES_PART].decode("utf-8")
    fonts = _section_items(styles_xml, "fonts", "font")
    fills = _section_items(styles_xml, "fills", "fill")
    borders = _section_items(styles_xml, "borders", "border")
    numfmts = _numfmts(styles_xml)
    cellxfs = _section_items(styles_xml, "cellXfs", "xf")

    resolved: dict[str, dict] = {"sheets": {}}
    for sheet_name, sheet_diff in diff.get("sheets", {}).items():
        part_name = SHEET_NAME_TO_PART.get(sheet_name)
        if part_name is None or part_name not in parts:
            continue
        sheet_xml = parts[part_name].decode("utf-8")
        cells = {}
        for ref in sheet_diff.get("cell_refs", []):
            style_index = _cell_style_index(sheet_xml, ref)
            if style_index is None or style_index >= len(cellxfs):
                continue
            cells[ref] = _resolve_xf(cellxfs[style_index], fonts, fills, borders, numfmts)
        resolved["sheets"][sheet_name] = {
            "cells": cells,
            "col_widths": sheet_diff.get("col_widths", {}),
            "extra_merges": sheet_diff.get("extra_merges", []),
        }
    return resolved


def capture_and_resolve(fixed_path: Path) -> dict:
    """Convenience wrapper: capture_style_diff + extract_cell_styles."""
    diff = capture_style_diff(fixed_path)
    return extract_cell_styles(fixed_path, diff)


def _find_or_append(items: list[str], entry: str) -> int:
    for index, existing in enumerate(items):
        if existing == entry:
            return index
    items.append(entry)
    return len(items) - 1


def _build_xf(xf_attrs: dict, font_id: int, fill_id: int, border_id: int, numfmt_id: int, extra_xml: str) -> str:
    attrs = dict(xf_attrs)
    attrs["fontId"] = str(font_id)
    attrs["fillId"] = str(fill_id)
    attrs["borderId"] = str(border_id)
    attrs["numFmtId"] = str(numfmt_id)
    attrs.setdefault("xfId", "0")
    attrs_str = " ".join(f'{key}="{value}"' for key, value in attrs.items())
    if extra_xml:
        return f"<xf {attrs_str}>{extra_xml}</xf>"
    return f"<xf {attrs_str}/>"


def _replace_section(styles_xml: str, tag: str, items: list[str]) -> str:
    inner = "".join(items)
    return re.sub(
        rf"<{tag}\b[^>]*>.*?</{tag}>",
        f'<{tag} count="{len(items)}">{inner}</{tag}>',
        styles_xml,
        count=1,
        flags=re.S,
    )


def _replace_numfmts(styles_xml: str, numfmts: dict[int, str]) -> str:
    inner = "".join(numfmts[k] for k in sorted(numfmts))
    new_section = f'<numFmts count="{len(numfmts)}">{inner}</numFmts>'
    if re.search(r"<numFmts\b[^>]*>.*?</numFmts>", styles_xml, re.S):
        return re.sub(r"<numFmts\b[^>]*>.*?</numFmts>", new_section, styles_xml, count=1, flags=re.S)
    # No existing numFmts section (template had none) -- insert right after
    # the opening <styleSheet ...> tag, since numFmts must come first.
    match = re.search(r"<styleSheet\b[^>]*>", styles_xml)
    return styles_xml[: match.end()] + new_section + styles_xml[match.end() :]


def _set_cell_style_index(sheet_xml: str, ref: str, new_index: int) -> str:
    pattern = re.compile(rf'<c r="{re.escape(ref)}"[^>]*>')
    match = pattern.search(sheet_xml)
    if not match:
        raise ValueError(f"cell {ref} not found")
    tag = match.group(0)
    self_close = tag.endswith("/>")
    inner = tag[:-2] if self_close else tag[:-1]
    inner = re.sub(r'\s*s="\d+"', "", inner)
    inner += f' s="{new_index}"'
    new_tag = inner + ("/>" if self_close else ">")
    return sheet_xml[: match.start()] + new_tag + sheet_xml[match.end() :]


def _col_letter_to_index(letter: str) -> int:
    index = 0
    for char in letter:
        index = index * 26 + (ord(char.upper()) - ord("A") + 1)
    return index


def _set_col_width(sheet_xml: str, letter: str, width: float) -> str:
    """Only touches single-column <col min=N max=N .../> entries, to avoid
    splitting a shared default-width range and risking unrelated columns.
    """
    col_index = _col_letter_to_index(letter)
    cols_match = re.search(r"<cols>(.*?)</cols>", sheet_xml, re.S)
    if not cols_match:
        new_cols = f'<cols><col min="{col_index}" max="{col_index}" width="{width}" customWidth="1"/></cols>'
        insert_at = sheet_xml.index("<sheetData")
        return sheet_xml[:insert_at] + new_cols + sheet_xml[insert_at:]

    inner = cols_match.group(1)
    entries = re.findall(r"<col\b[^>]*/>", inner)
    replaced = False
    new_entries = []
    for entry in entries:
        min_match = re.search(r'min="(\d+)"', entry)
        max_match = re.search(r'max="(\d+)"', entry)
        if (
            min_match
            and max_match
            and int(min_match.group(1)) == col_index
            and int(max_match.group(1)) == col_index
        ):
            new_entries.append(f'<col min="{col_index}" max="{col_index}" width="{width}" customWidth="1"/>')
            replaced = True
        else:
            new_entries.append(entry)
    if not replaced:
        new_entries.append(f'<col min="{col_index}" max="{col_index}" width="{width}" customWidth="1"/>')
    new_inner = "".join(new_entries)
    return sheet_xml[: cols_match.start()] + f"<cols>{new_inner}</cols>" + sheet_xml[cols_match.end() :]


def _add_merge(sheet_xml: str, merge_range: str) -> str:
    if f'ref="{merge_range}"' in sheet_xml:
        return sheet_xml
    match = re.search(r"<mergeCells\b[^>]*>(.*?)</mergeCells>", sheet_xml, re.S)
    if match:
        count_match = re.search(r'<mergeCells count="(\d+)"', sheet_xml)
        new_count = int(count_match.group(1)) + 1 if count_match else 1
        new_inner = match.group(1) + f'<mergeCell ref="{merge_range}"/>'
        new_section = f'<mergeCells count="{new_count}">{new_inner}</mergeCells>'
        return sheet_xml[: match.start()] + new_section + sheet_xml[match.end() :]
    # No existing mergeCells section: insert right after </sheetData>.
    insert_after = sheet_xml.index("</sheetData>") + len("</sheetData>")
    new_section = f'<mergeCells count="1"><mergeCell ref="{merge_range}"/></mergeCells>'
    return sheet_xml[:insert_after] + new_section + sheet_xml[insert_after:]


def apply_style_diff(workbook_path: Path, resolved_diff: dict) -> None:
    """Apply a previously-captured, resolved style diff (see
    extract_cell_styles) onto a freshly-generated workbook. Best-effort per
    cell: a saved reference that no longer matches the current sheet layout
    is skipped rather than raising, since this is a cosmetic enhancement on
    top of document generation, not a required step.
    """
    workbook_path = Path(workbook_path)
    parts = xl.load_zip_parts(workbook_path)
    styles_xml = parts[STYLES_PART].decode("utf-8")
    fonts = _section_items(styles_xml, "fonts", "font")
    fills = _section_items(styles_xml, "fills", "fill")
    borders = _section_items(styles_xml, "borders", "border")
    numfmts = _numfmts(styles_xml)
    cellxfs = _section_items(styles_xml, "cellXfs", "xf")

    styles_changed = False
    for sheet_name, sheet_diff in resolved_diff.get("sheets", {}).items():
        part_name = SHEET_NAME_TO_PART.get(sheet_name)
        if part_name is None or part_name not in parts:
            continue
        sheet_xml = parts[part_name].decode("utf-8")

        for ref, style in sheet_diff.get("cells", {}).items():
            try:
                font_id = _find_or_append(fonts, style["font_xml"])
                fill_id = _find_or_append(fills, style["fill_xml"])
                border_id = _find_or_append(borders, style["border_xml"])
                numfmt_id = style["numfmt_id"]
                if style.get("numfmt_xml") and numfmt_id not in numfmts:
                    numfmts[numfmt_id] = style["numfmt_xml"]
                xf_xml = _build_xf(style["xf_attrs"], font_id, fill_id, border_id, numfmt_id, style["xf_extra_xml"])
                xf_index = _find_or_append(cellxfs, xf_xml)
                sheet_xml = _set_cell_style_index(sheet_xml, ref, xf_index)
                styles_changed = True
            except (KeyError, ValueError):
                continue

        for letter, width in sheet_diff.get("col_widths", {}).items():
            try:
                sheet_xml = _set_col_width(sheet_xml, letter, width)
            except Exception:
                continue

        for merge_range in sheet_diff.get("extra_merges", []):
            try:
                sheet_xml = _add_merge(sheet_xml, merge_range)
            except Exception:
                continue

        parts[part_name] = sheet_xml.encode("utf-8")

    if styles_changed:
        styles_xml = _replace_section(styles_xml, "fonts", fonts)
        styles_xml = _replace_section(styles_xml, "fills", fills)
        styles_xml = _replace_section(styles_xml, "borders", borders)
        styles_xml = _replace_section(styles_xml, "cellXfs", cellxfs)
        styles_xml = _replace_numfmts(styles_xml, numfmts)
        parts[STYLES_PART] = styles_xml.encode("utf-8")

    xl.write_zip_parts(workbook_path, parts)
